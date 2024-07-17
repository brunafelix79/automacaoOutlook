import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import os
from openpyxl import load_workbook

# Configurações do servidor SMTP do Outlook.com
smtp_server = 'smtp-mail.outlook.com'
smtp_port = 587  # Porta TLS

# Informações de login
email_usuario = ''
senha = ''

# Caminho do arquivo Excel
caminho_arquivo_excel = r''

# Função para obter os destinatários e os caminhos dos anexos dinamicamente
def obter_destinatarios_e_anexos():
    wb = load_workbook(caminho_arquivo_excel)
    planilha = wb.active
    destinatarios_e_anexos = []

    # Iterar pelas linhas da coluna "EMAIL" e "NOME"
    for linha in planilha.iter_rows(min_row=2, max_col=6, values_only=True):
        email = linha[3]  # Coluna "EMAIL"
        nome = linha[1]  # Coluna "NOME"
        nome_completo = linha[0] #coluna "NOME COMPLETO"
        if email and nome:  # Verificar se há um e-mail e nome válidos
            caminho_arquivo_pdf = f'C:\\Users\\bruna\\Downloads\\{nome_completo}.pdf'  # Corrigir o formato do caminho do PDF
            if os.path.exists(caminho_arquivo_pdf):  # Verificar se o arquivo existe
                destinatarios_e_anexos.append((email, nome, caminho_arquivo_pdf))
            else:
                print(f"Arquivo não encontrado: {caminho_arquivo_pdf}")

    return destinatarios_e_anexos

# Função para enviar o e-mail
def enviar_email(destinatario, nome, caminho_arquivo_anexo):
    msg = MIMEMultipart()
    msg['From'] = email_usuario
    msg['To'] = destinatario
    msg['Subject'] = 'Assunto do E-mail'

    # Corpo do e-mail em texto plano
    body = f"""
    Olá, {nome}!

    Obrigada pela visita ao meu github
    
    Abraços, 
    """
    msg.attach(MIMEText(body, 'plain'))

    # Anexar arquivo PDF
    with open(caminho_arquivo_anexo, 'rb') as anexo:
        part = MIMEBase('application', 'pdf')
        part.set_payload(anexo.read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', f'attachment; filename="{os.path.basename(caminho_arquivo_anexo)}"')
    msg.attach(part)

    # Conexão com o servidor SMTP
    try:
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(email_usuario, senha)
            server.sendmail(email_usuario, destinatario, msg.as_string())
        print(f"E-mail enviado com sucesso para {destinatario}!")
    except Exception as e:
        print(f"Erro ao enviar e-mail para {destinatario}: {str(e)}")

# Chamada das funções principais
if __name__ == "__main__":
    destinatarios_e_anexos = obter_destinatarios_e_anexos()
    for destinatario, nome, caminho_anexo_pdf in destinatarios_e_anexos:
        enviar_email(destinatario, nome, caminho_anexo_pdf)
